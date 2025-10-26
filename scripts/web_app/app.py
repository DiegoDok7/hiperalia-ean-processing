import sys
import logging

# Configurar logging detallado
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

logger.info("🚀 Iniciando importaciones...")

from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for, Response, stream_with_context
logger.info("✓ Flask importado")

import os
import json
import requests
from datetime import datetime
logger.info("✓ Módulos estándar importados")

from PIL import Image
from io import BytesIO
import base64
logger.info("✓ PIL y IO importados")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
logger.info("✓ openpyxl importado")

from dotenv import load_dotenv
import tempfile
import shutil
import zipfile
import time
import re
logger.info("✓ Utilidades importadas")

# Cargar variables de entorno
load_dotenv()
logger.info("✓ Variables de entorno cargadas")

logger.info("🔧 Creando aplicación Flask...")
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'tu_clave_secreta_aqui')
logger.info("✓ Aplicación Flask creada")

# Configuración de directorios
logger.info("📁 Configurando directorios...")
# En producción usar /tmp, en desarrollo usar static/
if os.environ.get('RENDER'):
    UPLOAD_FOLDER = '/tmp/uploads'
    PROCESSED_FOLDER = '/tmp/processed'
    logger.info(f"🌐 Modo producción (Render) - usando /tmp")
else:
    UPLOAD_FOLDER = 'static/uploads'
    PROCESSED_FOLDER = 'static/processed'
    logger.info(f"💻 Modo desarrollo - usando static/")

# Crear directorios de forma segura
try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    logger.info(f"✓ Directorios creados: {UPLOAD_FOLDER}, {PROCESSED_FOLDER}")
except Exception as e:
    logger.warning(f"⚠️ No se pudieron crear directorios: {e}")
    # Continuar sin fallar, usar /tmp como fallback
    UPLOAD_FOLDER = '/tmp'
    PROCESSED_FOLDER = '/tmp'
    logger.info(f"✓ Usando fallback: /tmp")

def get_product_data(ean):
    """Obtiene datos del producto usando Open Food Facts API v2"""
    try:
        print(f"🔍 get_product_data recibió EAN: '{ean}' (tipo: {type(ean)})")  # Debug
        
        # Validar formato del EAN
        if not ean or not ean.isdigit() or len(ean) < 8 or len(ean) > 14:
            print(f"❌ EAN inválido: '{ean}' (longitud: {len(ean) if ean else 0})")  # Debug
            return {
                'success': False, 
                'error': 'El código EAN debe ser numérico y tener entre 8 y 14 dígitos'
            }
        
        # Usar API v2 con headers obligatorios
        url = f"https://world.openfoodfacts.org/api/v2/product/{ean}.json"
        headers = {
            "User-Agent": "MiApp/1.0 (miemail@example.com)"
        }
        
        print(f"🔍 Consultando API para EAN: {ean}")  # Debug
        print(f"🔍 URL: {url}")  # Debug
        response = requests.get(url, headers=headers, timeout=15)
        print(f"Status Code: {response.status_code}")  # Debug
        
        if response.status_code == 200:
            data = response.json()
            print(f"JSON Status: {data.get('status')}")  # Debug
            
            if data.get('status') == 1:
                product = data.get('product', {})
                print(f"Producto encontrado: {product.get('product_name', 'N/A')}")  # Debug
                
                return {
                    'success': True,
                    'data': {
                        'ean': ean,
                        'name': product.get('product_name', 'No disponible'),
                        'brand': product.get('brands', 'No disponible'),
                        'description': product.get('generic_name', 'No disponible'),
                        'category': product.get('categories', 'No disponible'),
                        'image_url': product.get('image_url', None),
                        'nutrition_grade': product.get('nutrition_grade_fr', 'No disponible'),
                        'ingredients': product.get('ingredients_text', 'No disponible'),
                        'allergens': product.get('allergens_tags', []),
                        'additives': product.get('additives_tags', []),
                        'nutriments': product.get('nutriments', {}),
                        'created_t': product.get('created_t', None),
                        'last_modified_t': product.get('last_modified_t', None)
                    }
                }
            else:
                # Producto no encontrado en la base de datos
                status_verbose = data.get('status_verbose', 'Producto no encontrado')
                return {
                    'success': False, 
                    'error': f'El producto con código EAN {ean} no se encuentra en nuestra base de datos. Verifica que el código sea correcto o intenta con otro producto.'
                }
        elif response.status_code == 404:
            return {
                'success': False, 
                'error': f'El producto con código EAN {ean} no existe en nuestra base de datos. Verifica que el código sea correcto.'
            }
        elif response.status_code == 429:
            return {
                'success': False, 
                'error': 'Demasiadas consultas. Por favor, espera un momento e intenta nuevamente.'
            }
        elif response.status_code >= 500:
            return {
                'success': False, 
                'error': 'Error del servidor. Por favor, intenta nuevamente en unos minutos.'
            }
        else:
            return {
                'success': False, 
                'error': f'Error de conexión (código {response.status_code}). Por favor, verifica tu conexión a internet e intenta nuevamente.'
            }
    except requests.exceptions.Timeout:
        return {
            'success': False, 
            'error': 'La consulta tardó demasiado tiempo. Por favor, verifica tu conexión a internet e intenta nuevamente.'
        }
    except requests.exceptions.ConnectionError:
        return {
            'success': False, 
            'error': 'No se pudo conectar al servidor. Por favor, verifica tu conexión a internet e intenta nuevamente.'
        }
    except Exception as e:
        return {
            'success': False, 
            'error': f'Error inesperado: {str(e)}. Por favor, intenta nuevamente.'
        }

def download_image(image_url, ean):
    """Descarga la imagen del producto en memoria (no guarda archivos)"""
    try:
        response = requests.get(image_url, timeout=15)
        if response.status_code == 200:
            # Convertir a base64 para mostrar en la web sin guardar archivo
            image_base64 = base64.b64encode(response.content).decode('utf-8')
            return {
                'success': True, 
                'image_data': image_base64,
                'content_type': 'image/jpeg',
                'size': len(response.content)
            }
        else:
            return {'success': False, 'error': f'Error descargando imagen: {response.status_code}'}
    except Exception as e:
        return {'success': False, 'error': f'Error descargando imagen: {str(e)}'}

def search_product_web_data(ean, product_name, api_key):
    """Busca información adicional del producto en internet usando Gemini 2.5 Flash-Lite"""
    try:
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent"
        headers = {
            "x-goog-api-key": api_key,
            "Content-Type": "application/json"
        }
        
        prompt = f"""
        Busca información detallada en internet sobre el producto con código EAN {ean} {f'y nombre "{product_name}"' if product_name and product_name != 'No disponible' else ''}.
        
        Proporciona la información en formato JSON con los siguientes campos (si no encuentras un campo, usa "No disponible"):
        {{
            "nombre": "nombre completo del producto",
            "descripcion": "descripción detallada del producto",
            "marca": "marca del producto",
            "categoria": "categoría principal",
            "categoria_path": "ruta completa de categorías (ej: Alimentación > Snacks > Galletas)",
            "departamento": "departamento al que pertenece",
            "producto_tipo": "tipo de producto",
            "ingredientes": "lista de ingredientes",
            "alergenos": "alérgenos presentes",
            "organico": "si/no",
            "no_gmo": "si/no",
            "altura": "altura en cm",
            "ancho": "ancho en cm",
            "largo": "largo en cm",
            "upc": "código UPC si está disponible",
            "precio_estimado": "precio estimado en euros"
        }}
        
        IMPORTANTE: Responde SOLO con el objeto JSON, sin texto adicional antes o después.
        """
        
        payload = {
            "contents": [{
                "parts": [{"text": prompt}]
            }],
            "generationConfig": {
                "temperature": 0.1,
                "topK": 40,
                "topP": 0.95,
                "maxOutputTokens": 2048,
            }
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        
        if response.status_code == 200:
            data = response.json()
            
            if "candidates" in data and data["candidates"]:
                candidate = data["candidates"][0]
                if "content" in candidate and "parts" in candidate["content"]:
                    for part in candidate["content"]["parts"]:
                        if "text" in part:
                            text_response = part["text"].strip()
                            # Intentar extraer JSON de la respuesta
                            try:
                                # Buscar el JSON en la respuesta
                                json_start = text_response.find('{')
                                json_end = text_response.rfind('}') + 1
                                if json_start != -1 and json_end > json_start:
                                    json_str = text_response[json_start:json_end]
                                    web_data = json.loads(json_str)
                                    return {
                                        'success': True,
                                        'data': web_data
                                    }
                                else:
                                    return {'success': False, 'error': 'No se encontró JSON en la respuesta'}
                            except json.JSONDecodeError:
                                return {'success': False, 'error': 'Error parseando JSON de la respuesta'}
            
            return {'success': False, 'error': 'No se encontró contenido en la respuesta'}
        else:
            return {'success': False, 'error': f'Error API Gemini: {response.status_code}'}
    
    except Exception as e:
        return {'success': False, 'error': f'Error buscando datos web: {str(e)}'}

def search_and_download_product_image(ean, product_name, image_url_fallback=None):
    """Busca y descarga imagen del producto desde múltiples fuentes"""
    try:
        # Lista de URLs a intentar
        urls_to_try = []
        
        # 1. Intentar con EAN lookup services
        urls_to_try.append({
            'url': f'https://images.openfoodfacts.org/images/products/{ean[:3]}/{ean[3:6]}/{ean[6:9]}/{ean[9:]}/front_en.jpg',
            'source': 'OpenFoodFacts (alta resolución)'
        })
        
        urls_to_try.append({
            'url': f'https://world.openfoodfacts.org/images/products/{ean[:3]}/{ean[3:6]}/{ean[6:9]}/{ean[9:]}/1.jpg',
            'source': 'OpenFoodFacts (imagen 1)'
        })
        
        # 2. Si hay URL fallback de OpenFoodFacts, usarla
        if image_url_fallback:
            urls_to_try.append({
                'url': image_url_fallback,
                'source': 'OpenFoodFacts (API)'
            })
        
        # 3. Intentar con servicios de imágenes de EAN
        urls_to_try.append({
            'url': f'https://www.ean-search.org/images/{ean}.jpg',
            'source': 'EAN-Search'
        })
        
        # Intentar descargar de cada URL
        for item in urls_to_try:
            try:
                logger.info(f"  🔍 Intentando descargar imagen desde: {item['source']}")
                img_response = requests.get(item['url'], timeout=10, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                })
                
                if img_response.status_code == 200 and len(img_response.content) > 1000:  # Mínimo 1KB
                    # Verificar que sea una imagen válida
                    try:
                        img = Image.open(BytesIO(img_response.content))
                        width, height = img.size
                        
                        # Preferir imágenes de tamaño razonable
                        if width >= 200 and height >= 200:
                            logger.info(f"  ✓ Imagen encontrada: {width}x{height} desde {item['source']}")
                            
                            # Convertir a base64
                            image_base64 = base64.b64encode(img_response.content).decode('utf-8')
                            
                            return {
                                'success': True,
                                'image_data': image_base64,
                                'content_type': img_response.headers.get('content-type', 'image/jpeg'),
                                'size': len(img_response.content),
                                'source': item['source'],
                                'quality': 'alta' if width >= 800 else 'media' if width >= 400 else 'baja'
                            }
                    except Exception as img_error:
                        logger.warning(f"  ⚠️ No es una imagen válida desde {item['source']}: {img_error}")
                        continue
                        
            except Exception as download_error:
                logger.warning(f"  ⚠️ Error descargando desde {item['source']}: {str(download_error)}")
                continue
        
        # Si no se encontró ninguna imagen
        return {
            'success': False,
            'error': 'No se pudo encontrar imagen del producto en ninguna fuente'
        }
    
    except Exception as e:
        return {'success': False, 'error': f'Error buscando imagen: {str(e)}'}

def enhance_image_with_gemini(image_data_base64, prompt, api_key):
    """Mejora la imagen usando Google Gemini API (trabaja en memoria)"""
    try:
        # Preparar payload para Gemini
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-image-preview:generateContent"
        headers = {
            "x-goog-api-key": api_key,
            "Content-Type": "application/json"
        }
        
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": prompt
                        },
                        {
                            "inline_data": {
                                "mime_type": "image/jpeg",
                                "data": image_data_base64
                            }
                        }
                    ]
                }
            ]
        }
        
        # Llamar a la API
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        
        if response.status_code == 200:
            data = response.json()
            
            if "candidates" in data and data["candidates"]:
                candidate = data["candidates"][0]
                if "content" in candidate and "parts" in candidate["content"]:
                    for part in candidate["content"]["parts"]:
                        if "inlineData" in part:
                            # Devolver imagen mejorada como base64
                            enhanced_image_base64 = part['inlineData']['data']
                            return {
                                'success': True, 
                                'image_data': enhanced_image_base64,
                                'content_type': 'image/png'
                            }
            
            return {'success': False, 'error': 'No se pudo procesar la imagen con IA'}
        else:
            return {'success': False, 'error': f'Error API Gemini: {response.status_code}'}
    
    except Exception as e:
        return {'success': False, 'error': f'Error procesando imagen: {str(e)}'}

def combine_product_data(ean, off_data, web_data):
    """Combina datos de OpenFoodFacts y búsqueda web para crear registro completo"""
    try:
        # Generar Product ID único
        product_id = f"PROD-{ean}"
        
        # Combinar datos priorizando la mejor fuente para cada campo
        combined = {
            # Campos para PrestaShop
            'Product ID': product_id,
            'Imagen': '',  # Se llenará con la ruta de la imagen procesada
            'Nombre': web_data.get('nombre', off_data.get('name', 'No disponible')),
            'Referencia': ean,
            'Categoría': web_data.get('categoria', off_data.get('category', 'No disponible')),
            'Precio (imp. excl.)': '',  # Dejar vacío según requerimientos
            'Precio (imp. incl.)': '',  # Dejar vacío según requerimientos
            'Cantidad': '0',  # Valor por defecto
            
            # Campos adicionales de interés
            'Codigo': ean,
            'Codigo Tipo': 'EAN',
            'Nombre Producto': web_data.get('nombre', off_data.get('name', 'No disponible')),
            'Descripcion': web_data.get('descripcion', off_data.get('description', 'No disponible')),
            'Marca': web_data.get('marca', off_data.get('brand', 'No disponible')),
            'Categoria': web_data.get('categoria', off_data.get('category', 'No disponible')),
            'Categoria Path': web_data.get('categoria_path', 'No disponible'),
            'Departamento': web_data.get('departamento', 'No disponible'),
            'Producto Tipo': web_data.get('producto_tipo', 'No disponible'),
            'Imagen Url': off_data.get('image_url', ''),
            'Upc': web_data.get('upc', 'No disponible'),
            'Ean': ean,
            'Ingredientes': web_data.get('ingredientes', off_data.get('ingredients', 'No disponible')),
            'Alergenos': web_data.get('alergenos', ', '.join(off_data.get('allergens', [])) if off_data.get('allergens') else 'No disponible'),
            'Organico': web_data.get('organico', 'No disponible'),
            'No Gmo': web_data.get('no_gmo', 'No disponible'),
            'Altura': web_data.get('altura', 'No disponible'),
            'Ancho': web_data.get('ancho', 'No disponible'),
            'Largo': web_data.get('largo', 'No disponible'),
            'Barcode Url': f'https://barcode.tec-it.com/barcode.ashx?data={ean}&code=EAN13',
            'Producto Encontrado': 'si'
        }
        
        return combined
    
    except Exception as e:
        logger.error(f"Error combinando datos: {e}")
        # Retornar estructura básica con datos mínimos
        return {
            'Product ID': f"PROD-{ean}",
            'Imagen': '',
            'Nombre': 'Error procesando datos',
            'Referencia': ean,
            'Categoría': 'No disponible',
            'Precio (imp. excl.)': '',
            'Precio (imp. incl.)': '',
            'Cantidad': '0',
            'Codigo': ean,
            'Codigo Tipo': 'EAN',
            'Nombre Producto': 'Error procesando datos',
            'Descripcion': 'No disponible',
            'Marca': 'No disponible',
            'Categoria': 'No disponible',
            'Categoria Path': 'No disponible',
            'Departamento': 'No disponible',
            'Producto Tipo': 'No disponible',
            'Imagen Url': '',
            'Upc': 'No disponible',
            'Ean': ean,
            'Ingredientes': 'No disponible',
            'Alergenos': 'No disponible',
            'Organico': 'No disponible',
            'No Gmo': 'No disponible',
            'Altura': 'No disponible',
            'Ancho': 'No disponible',
            'Largo': 'No disponible',
            'Barcode Url': f'https://barcode.tec-it.com/barcode.ashx?data={ean}&code=EAN13',
            'Producto Encontrado': 'no'
        }

def remove_white_background(image_data_base64):
    """Remueve el fondo blanco de una imagen usando rembg"""
    try:
        # Importación lazy de rembg para evitar timeout en el inicio
        try:
            from rembg import remove
        except ImportError as ie:
            print(f"⚠️ rembg no disponible: {ie}")
            return {
                'success': False,
                'error': 'Librería rembg no disponible en este entorno'
            }
        
        # Decodificar la imagen base64
        image_bytes = base64.b64decode(image_data_base64)
        
        # Abrir imagen con PIL
        input_image = Image.open(BytesIO(image_bytes))
        
        # Remover fondo usando rembg
        output_image = remove(input_image)
        
        # Convertir a base64
        output_buffer = BytesIO()
        output_image.save(output_buffer, format='PNG')
        output_bytes = output_buffer.getvalue()
        output_base64 = base64.b64encode(output_bytes).decode('utf-8')
        
        return {
            'success': True,
            'image_data': output_base64,
            'content_type': 'image/png'
        }
    
    except Exception as e:
        return {'success': False, 'error': f'Error removiendo fondo: {str(e)}'}

def create_excel_data(product_data, ean):
    """Crea datos Excel en memoria (no guarda archivos)"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos del Producto"
        
        # Encabezados
        headers = [
            'Campo', 'Valor'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos del producto
        data_rows = [
            ('EAN', product_data['ean']),
            ('Nombre', product_data['name']),
            ('Marca', product_data['brand']),
            ('Descripción', product_data['description']),
            ('Categoría', product_data['category']),
            ('Grado Nutricional', product_data['nutrition_grade']),
            ('Ingredientes', product_data['ingredients']),
            ('Alérgenos', ', '.join(product_data['allergens']) if product_data['allergens'] else 'Ninguno'),
            ('Aditivos', ', '.join(product_data['additives']) if product_data['additives'] else 'Ninguno'),
            ('Fecha Creación', datetime.fromtimestamp(product_data['created_t']).strftime('%Y-%m-%d %H:%M:%S') if product_data['created_t'] else 'No disponible'),
            ('Última Modificación', datetime.fromtimestamp(product_data['last_modified_t']).strftime('%Y-%m-%d %H:%M:%S') if product_data['last_modified_t'] else 'No disponible')
        ]
        
        for row, (field, value) in enumerate(data_rows, 2):
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=value)
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        
        # Guardar en memoria como bytes
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_data = excel_buffer.getvalue()
        excel_base64 = base64.b64encode(excel_data).decode('utf-8')
        
        return {
            'success': True, 
            'excel_data': excel_base64,
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'size': len(excel_data)
        }
    
    except Exception as e:
        return {'success': False, 'error': f'Error creando Excel: {str(e)}'}

# Health check endpoint para Render
@app.route('/health')
def health():
    """Endpoint de health check para monitoreo"""
    return jsonify({
        'status': 'healthy',
        'service': 'ean-automation',
        'timestamp': datetime.now().isoformat()
    }), 200

@app.route('/')
def index():
    """Pantalla de inicio con opciones de búsqueda"""
    return render_template('index.html')

@app.route('/search_individual')
def search_individual():
    """Pantalla de búsqueda individual"""
    return render_template('search_individual.html')

@app.route('/search_bulk')
def search_bulk():
    """Pantalla de búsqueda por grupos"""
    return render_template('search_bulk.html')

@app.route('/images_only')
def images_only():
    """Pantalla de procesamiento solo de imágenes"""
    return render_template('images_only.html')

@app.route('/process_ean', methods=['POST'])
def process_ean():
    try:
        ean = request.form.get('ean', '').strip()
        print(f"🔍 EAN recibido en process_ean: '{ean}'")  # Debug
        
        if not ean:
            return jsonify({'success': False, 'error': 'EAN requerido'})
        
        # Obtener datos del producto
        print(f"🔍 Llamando get_product_data con EAN: '{ean}'")  # Debug
        product_result = get_product_data(ean)
        
        if not product_result['success']:
            return jsonify(product_result)
        
        product_data = product_result['data']
        result = {
            'success': True,
            'product_data': product_data,
            'images': {},
            'files': {}
        }
        
        # Buscar y descargar imagen desde múltiples fuentes
        api_key = os.getenv("GEMINI_API_KEY")
        # Buscar imagen de alta calidad desde múltiples fuentes
        image_search_result = search_and_download_product_image(
            ean, 
            product_data.get('name', 'No disponible'),
            product_data.get('image_url')  # URL de OpenFoodFacts como fallback
        )
        
        if image_search_result['success']:
            result['images']['original'] = {
                'data': image_search_result['image_data'],
                'content_type': image_search_result['content_type'],
                'size': image_search_result['size'],
                'source': image_search_result.get('source', 'internet'),
                'quality': image_search_result.get('quality', 'desconocida')
            }
            
            # Mejorar imagen con IA si hay API key
            if api_key:
                prompt = ("Take the provided product image and enhance it for PrestaShop e-commerce platform. "
                        "Create a square image (800x800 pixels) with these specifications: "
                        "1. Remove the background completely and replace it with pure white (#FFFFFF). "
                        "2. Center the product perfectly in the frame. "
                        "3. The product should occupy 80-85% of the image space, leaving appropriate margins. "
                        "4. Show the product from the front in its most recognizable angle. "
                        "5. Enhance lighting to be even and professional, eliminating shadows on the background. "
                        "6. Improve sharpness and color accuracy for high-quality zoom capability. "
                        "7. Ensure the product looks professional, clean, and appealing for online sales. "
                        "8. Keep the product realistic and true to its original appearance. "
                        "The final image must be optimized for PrestaShop product listings with consistent quality.")
                
                enhance_result = enhance_image_with_gemini(image_search_result['image_data'], prompt, api_key)
                if enhance_result['success']:
                    # Remover fondo blanco usando rembg
                    remove_bg_result = remove_white_background(enhance_result['image_data'])
                    if remove_bg_result['success']:
                        result['images']['enhanced'] = {
                            'data': remove_bg_result['image_data'],
                            'content_type': remove_bg_result['content_type']
                        }
                    else:
                        # Si falla la remoción de fondo, usar la imagen mejorada con IA
                        result['images']['enhanced'] = {
                            'data': enhance_result['image_data'],
                            'content_type': enhance_result['content_type']
                        }
                        result['bg_removal_warning'] = remove_bg_result['error']
                else:
                    result['ai_error'] = enhance_result['error']
        else:
            result['image_search_error'] = image_search_result['error']
        
        # Crear datos Excel
        excel_result = create_excel_data(product_data, ean)
        if excel_result['success']:
            result['files']['excel'] = {
                'data': excel_result['excel_data'],
                'content_type': excel_result['content_type'],
                'size': excel_result['size']
            }
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error procesando EAN: {str(e)}'})

def sanitize_filename(name):
    """Sanitiza un nombre para usarlo como nombre de archivo"""
    if not name or name == 'No disponible':
        return 'sin_nombre'
    # Remover caracteres no válidos
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    # Limitar longitud
    name = name[:50]
    # Remover espacios al inicio y final
    name = name.strip()
    # Reemplazar espacios con guiones bajos
    name = name.replace(' ', '_')
    return name if name else 'sin_nombre'

def create_bulk_excel(products_data):
    """Crea un Excel con múltiples productos organizados por columnas para PrestaShop"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos para PrestaShop"
        
        # Encabezados según especificación
        headers = [
            # Campos principales para PrestaShop
            'Product ID', 'Imagen', 'Nombre', 'Referencia', 'Categoría',
            'Precio (imp. excl.)', 'Precio (imp. incl.)', 'Cantidad',
            # Campos adicionales de interés
            'Codigo', 'Codigo Tipo', 'Nombre Producto', 'Descripcion', 'Marca',
            'Categoria', 'Categoria Path', 'Departamento', 'Producto Tipo',
            'Imagen Url', 'Upc', 'Ean', 'Ingredientes', 'Alergenos',
            'Organico', 'No Gmo', 'Altura', 'Ancho', 'Largo',
            'Barcode Url', 'Producto Encontrado'
        ]
        
        # Estilo para encabezados principales (PrestaShop)
        header_fill_main = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_main = Font(bold=True, color="FFFFFF", size=11)
        
        # Estilo para encabezados adicionales
        header_fill_additional = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font_additional = Font(bold=True, color="FFFFFF", size=10)
        
        # Aplicar encabezados con estilos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Primeros 8 campos son para PrestaShop
            if col <= 8:
                cell.fill = header_fill_main
                cell.font = header_font_main
            else:
                cell.fill = header_fill_additional
                cell.font = header_font_additional
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Datos de productos
        for row_idx, product in enumerate(products_data, 2):
            # Campos para PrestaShop
            ws.cell(row=row_idx, column=1, value=product.get('Product ID', ''))
            ws.cell(row=row_idx, column=2, value=product.get('Imagen', ''))
            ws.cell(row=row_idx, column=3, value=product.get('Nombre', 'No disponible'))
            ws.cell(row=row_idx, column=4, value=product.get('Referencia', ''))
            ws.cell(row=row_idx, column=5, value=product.get('Categoría', 'No disponible'))
            ws.cell(row=row_idx, column=6, value=product.get('Precio (imp. excl.)', ''))
            ws.cell(row=row_idx, column=7, value=product.get('Precio (imp. incl.)', ''))
            ws.cell(row=row_idx, column=8, value=product.get('Cantidad', '0'))
            
            # Campos adicionales de interés
            ws.cell(row=row_idx, column=9, value=product.get('Codigo', ''))
            ws.cell(row=row_idx, column=10, value=product.get('Codigo Tipo', 'EAN'))
            ws.cell(row=row_idx, column=11, value=product.get('Nombre Producto', 'No disponible'))
            ws.cell(row=row_idx, column=12, value=product.get('Descripcion', 'No disponible'))
            ws.cell(row=row_idx, column=13, value=product.get('Marca', 'No disponible'))
            ws.cell(row=row_idx, column=14, value=product.get('Categoria', 'No disponible'))
            ws.cell(row=row_idx, column=15, value=product.get('Categoria Path', 'No disponible'))
            ws.cell(row=row_idx, column=16, value=product.get('Departamento', 'No disponible'))
            ws.cell(row=row_idx, column=17, value=product.get('Producto Tipo', 'No disponible'))
            ws.cell(row=row_idx, column=18, value=product.get('Imagen Url', ''))
            ws.cell(row=row_idx, column=19, value=product.get('Upc', 'No disponible'))
            ws.cell(row=row_idx, column=20, value=product.get('Ean', ''))
            ws.cell(row=row_idx, column=21, value=product.get('Ingredientes', 'No disponible'))
            ws.cell(row=row_idx, column=22, value=product.get('Alergenos', 'No disponible'))
            ws.cell(row=row_idx, column=23, value=product.get('Organico', 'No disponible'))
            ws.cell(row=row_idx, column=24, value=product.get('No Gmo', 'No disponible'))
            ws.cell(row=row_idx, column=25, value=product.get('Altura', 'No disponible'))
            ws.cell(row=row_idx, column=26, value=product.get('Ancho', 'No disponible'))
            ws.cell(row=row_idx, column=27, value=product.get('Largo', 'No disponible'))
            ws.cell(row=row_idx, column=28, value=product.get('Barcode Url', ''))
            ws.cell(row=row_idx, column=29, value=product.get('Producto Encontrado', 'no'))
        
        # Ajustar ancho de columnas
        column_widths = {
            1: 15,  # Product ID
            2: 30,  # Imagen
            3: 35,  # Nombre
            4: 15,  # Referencia
            5: 25,  # Categoría
            6: 15,  # Precio excl
            7: 15,  # Precio incl
            8: 10,  # Cantidad
            9: 15,  # Codigo
            10: 12, # Codigo Tipo
            11: 35, # Nombre Producto
            12: 50, # Descripcion
            13: 20, # Marca
            14: 25, # Categoria
            15: 40, # Categoria Path
            16: 20, # Departamento
            17: 20, # Producto Tipo
            18: 40, # Imagen Url
            19: 15, # Upc
            20: 15, # Ean
            21: 50, # Ingredientes
            22: 30, # Alergenos
            23: 10, # Organico
            24: 10, # No Gmo
            25: 10, # Altura
            26: 10, # Ancho
            27: 10, # Largo
            28: 40, # Barcode Url
            29: 15  # Producto Encontrado
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_data = excel_buffer.getvalue()
        
        return excel_data
    
    except Exception as e:
        logger.error(f"Error creando Excel bulk: {str(e)}")
        return None

@app.route('/process_bulk', methods=['POST'])
def process_bulk():
    """Procesa múltiples EANs y genera un ZIP con Excel e imágenes"""
    def generate():
        try:
            logger.info("📦 Iniciando process_bulk...")
            eans_json = request.form.get('eans', '[]')
            eans = json.loads(eans_json)
            logger.info(f"📊 Cantidad de EANs recibidos: {len(eans)}")
            
            if not eans:
                logger.warning("⚠️ No se recibieron códigos EAN")
                yield f"data: {json.dumps({'type': 'error', 'message': 'No se recibieron códigos EAN'})}\n\n"
                return
            
            # Limitar cantidad de EANs para evitar timeout
            max_eans = 50
            if len(eans) > max_eans:
                logger.warning(f"⚠️ Limitando procesamiento a {max_eans} EANs")
                yield f"data: {json.dumps({'type': 'warning', 'message': f'Se procesarán solo los primeros {max_eans} EANs de {len(eans)}'})}\n\n"
                eans = eans[:max_eans]
            
            products_data = []
            images_data = []
            api_key = os.getenv("GEMINI_API_KEY")
            
            # Procesar cada EAN
            for idx, ean in enumerate(eans):
                logger.info(f"🔄 Procesando EAN {idx+1}/{len(eans)}: {ean}")
                ean = ean.strip()
                
                # 1. Obtener datos de OpenFoodFacts
                try:
                    product_result = get_product_data(ean)
                    logger.info(f"  ✓ Datos OFF obtenidos para {ean}: {product_result.get('success', False)}")
                except Exception as e:
                    logger.error(f"  ❌ Error obteniendo datos OFF para {ean}: {e}")
                    yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': False, 'message': f'Error: {str(e)}'})}\n\n"
                    continue
                
                if product_result['success']:
                    off_product = product_result['data']
                    
                    # 2. Buscar datos adicionales con Gemini Web Search
                    web_data = {}
                    if api_key:
                        logger.info(f"  🌐 Buscando datos web con Gemini para {ean}")
                        try:
                            web_result = search_product_web_data(
                                ean, 
                                off_product.get('name', ''), 
                                api_key
                            )
                            if web_result['success']:
                                web_data = web_result['data']
                                logger.info(f"  ✓ Datos web obtenidos para {ean}")
                            else:
                                logger.warning(f"  ⚠️ No se pudieron obtener datos web: {web_result.get('error', 'Unknown')}")
                        except Exception as e:
                            logger.error(f"  ❌ Error en búsqueda web para {ean}: {e}")
                    
                    # 3. Combinar datos de OpenFoodFacts + Gemini Web
                    combined_product = combine_product_data(ean, off_product, web_data)
                    
                    # 4. Buscar y procesar imagen desde múltiples fuentes
                    image_filename = ''
                    logger.info(f"  🖼️ Buscando imagen para {ean}")
                    try:
                        # Buscar imagen de alta calidad desde múltiples fuentes
                        image_search_result = search_and_download_product_image(
                            ean,
                            off_product.get('name', 'No disponible'),
                            off_product.get('image_url')  # URL de OpenFoodFacts como fallback
                        )
                        
                        if image_search_result['success']:
                            logger.info(f"  ✓ Imagen encontrada (fuente: {image_search_result.get('source', 'desconocida')})")
                            
                            # Mejorar imagen con Gemini Image Preview si hay API key
                            if api_key:
                                logger.info(f"  🤖 Mejorando imagen con IA para {ean}")
                                prompt = ("Take the provided product image and enhance it for PrestaShop e-commerce platform. "
                                        "Create a square image (800x800 pixels) with these specifications: "
                                        "1. Remove the background completely and replace it with pure white (#FFFFFF). "
                                        "2. Center the product perfectly in the frame. "
                                        "3. The product should occupy 80-85% of the image space, leaving appropriate margins. "
                                        "4. Show the product from the front in its most recognizable angle. "
                                        "5. Enhance lighting to be even and professional, eliminating shadows on the background. "
                                        "6. Improve sharpness and color accuracy for high-quality zoom capability. "
                                        "7. Ensure the product looks professional, clean, and appealing for online sales. "
                                        "8. Keep the product realistic and true to its original appearance. "
                                        "The final image must be optimized for PrestaShop product listings with consistent quality.")
                                
                                try:
                                    enhance_result = enhance_image_with_gemini(image_search_result['image_data'], prompt, api_key)
                                    if enhance_result['success']:
                                        logger.info(f"  ✓ Imagen mejorada con IA para {ean}")
                                        # Remover fondo con rembg
                                        logger.info(f"  🎨 Removiendo fondo para {ean}")
                                        try:
                                            remove_bg_result = remove_white_background(enhance_result['image_data'])
                                            if remove_bg_result['success']:
                                                image_data_final = remove_bg_result['image_data']
                                            else:
                                                image_data_final = enhance_result['image_data']
                                        except Exception as e:
                                            logger.error(f"  ❌ Error en rembg para {ean}: {e}")
                                            image_data_final = enhance_result['image_data']
                                        
                                        # Guardar imagen - Solo EAN como nombre
                                        image_filename = f"{ean}.png"
                                        
                                        images_data.append({
                                            'filename': image_filename,
                                            'data': image_data_final
                                        })
                                        logger.info(f"  ✓ Imagen guardada: {image_filename}")
                                    else:
                                        logger.warning(f"  ⚠️ Error mejorando imagen: {enhance_result.get('error', 'Unknown')}")
                                except Exception as e:
                                    logger.error(f"  ❌ Excepción en mejora de imagen para {ean}: {e}")
                            else:
                                # Si no hay API key, usar imagen original - Solo EAN como nombre
                                image_filename = f"{ean}.png"
                                
                                images_data.append({
                                    'filename': image_filename,
                                    'data': image_search_result['image_data']
                                })
                                logger.info(f"  ✓ Imagen guardada (sin IA): {image_filename}")
                        else:
                            logger.warning(f"  ⚠️ No se pudo encontrar imagen: {image_search_result.get('error', 'Unknown')}")
                    except Exception as e:
                        logger.error(f"  ❌ Error buscando/procesando imagen para {ean}: {e}")
                    
                    # Actualizar ruta de imagen en datos combinados
                    if image_filename:
                        combined_product['Imagen'] = f"imagenes/{image_filename}"
                    
                    products_data.append(combined_product)
                    yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': True, 'message': 'Procesado correctamente'})}\n\n"
                    
                else:
                    # Si falla OpenFoodFacts, crear registro básico marcado como no encontrado
                    error_msg = product_result.get('error', 'Error desconocido')
                    logger.warning(f"  ⚠️ Producto no encontrado en OFF: {ean}")
                    
                    # Intentar buscar solo con Gemini
                    web_data = {}
                    if api_key:
                        try:
                            web_result = search_product_web_data(ean, '', api_key)
                            if web_result['success']:
                                web_data = web_result['data']
                        except Exception as e:
                            logger.error(f"  ❌ Error en búsqueda web alternativa para {ean}: {e}")
                    
                    combined_product = combine_product_data(ean, {}, web_data)
                    combined_product['Producto Encontrado'] = 'no'
                    products_data.append(combined_product)
                    
                    yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': False, 'message': f'No encontrado en OFF, datos web agregados'})}\n\n"
            
            # Crear archivo ZIP con Excel e imágenes
            logger.info(f"📦 Creando ZIP final con {len(products_data)} productos y {len(images_data)} imágenes")
            if products_data:
                try:
                    # Verificar estructura de products_data
                    logger.info(f"  🔍 Primer producto ejemplo: {list(products_data[0].keys()) if products_data else 'vacío'}")
                    
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        # Agregar Excel
                        logger.info("  📊 Creando Excel...")
                        excel_data = create_bulk_excel(products_data)
                        if excel_data:
                            zip_file.writestr('productos_prestashop.xlsx', excel_data)
                            logger.info(f"  ✓ Excel agregado ({len(excel_data)} bytes)")
                        else:
                            logger.error("  ❌ Excel data es None!")
                        
                        # Agregar imágenes en una carpeta
                        logger.info(f"  🖼️ Agregando {len(images_data)} imágenes...")
                        for img in images_data:
                            zip_file.writestr(f"imagenes/{img['filename']}", base64.b64decode(img['data']))
                        logger.info("  ✓ Imágenes agregadas")
                        
                        # Listar contenido del ZIP
                        zip_contents = zip_file.namelist()
                        logger.info(f"  📋 Contenido del ZIP: {zip_contents}")
                    
                    zip_data = zip_buffer.getvalue()
                    zip_base64 = base64.b64encode(zip_data).decode('utf-8')
                    logger.info(f"✅ ZIP creado exitosamente ({len(zip_data)} bytes, {len(zip_base64)} base64)")
                    
                    yield f"data: {json.dumps({'type': 'complete', 'zip_data': zip_base64})}\n\n"
                except Exception as e:
                    logger.error(f"❌ Error creando ZIP: {e}", exc_info=True)
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Error creando ZIP: {str(e)}'})}\n\n"
            else:
                logger.warning("⚠️ No hay productos para procesar")
                yield f"data: {json.dumps({'type': 'error', 'message': 'No se pudieron procesar productos'})}\n\n"
        
        except Exception as e:
            logger.error(f"❌ ERROR FATAL en process_bulk: {e}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'message': f'Error: {str(e)}'})}\n\n"
    
    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/process_images_only', methods=['POST'])
def process_images_only():
    """Procesa solo imágenes de múltiples EANs sin procesar datos de productos"""
    def generate():
        try:
            logger.info("🖼️ Iniciando process_images_only...")
            eans_json = request.form.get('eans', '[]')
            eans = json.loads(eans_json)
            logger.info(f"📊 Cantidad de EANs recibidos: {len(eans)}")
            
            if not eans:
                logger.warning("⚠️ No se recibieron códigos EAN")
                yield f"data: {json.dumps({'type': 'error', 'message': 'No se recibieron códigos EAN'})}\n\n"
                return
            
            # Limitar cantidad de EANs para evitar timeout
            max_eans = 50
            if len(eans) > max_eans:
                logger.warning(f"⚠️ Limitando procesamiento a {max_eans} EANs")
                yield f"data: {json.dumps({'type': 'warning', 'message': f'Se procesarán solo los primeros {max_eans} EANs de {len(eans)}'})}\n\n"
                eans = eans[:max_eans]
            
            images_data = []
            api_key = os.getenv("GEMINI_API_KEY")
            
            # Procesar cada EAN - SOLO IMÁGENES
            for idx, ean in enumerate(eans):
                logger.info(f"🔄 Procesando imagen {idx+1}/{len(eans)}: {ean}")
                ean = ean.strip()
                
                try:
                    # Obtener solo datos básicos de OFF para el nombre (sin procesar con Gemini)
                    product_result = get_product_data(ean)
                    product_name = 'producto'
                    image_url_fallback = None
                    
                    if product_result['success']:
                        off_product = product_result['data']
                        product_name = off_product.get('name', 'producto')
                        image_url_fallback = off_product.get('image_url')
                    
                    # Buscar imagen desde múltiples fuentes
                    logger.info(f"  🖼️ Buscando imagen para {ean}")
                    image_search_result = search_and_download_product_image(
                        ean,
                        product_name,
                        image_url_fallback
                    )
                    
                    if image_search_result['success']:
                        logger.info(f"  ✓ Imagen encontrada (fuente: {image_search_result.get('source', 'desconocida')})")
                        
                        # Mejorar imagen con Gemini Image Preview si hay API key
                        if api_key:
                            logger.info(f"  🤖 Mejorando imagen con IA para {ean}")
                            prompt = ("Take the provided product image and enhance it for PrestaShop e-commerce platform. "
                                    "Create a square image (800x800 pixels) with these specifications: "
                                    "1. Remove the background completely and replace it with pure white (#FFFFFF). "
                                    "2. Center the product perfectly in the frame. "
                                    "3. The product should occupy 80-85% of the image space, leaving appropriate margins. "
                                    "4. Show the product from the front in its most recognizable angle. "
                                    "5. Enhance lighting to be even and professional, eliminating shadows on the background. "
                                    "6. Improve sharpness and color accuracy for high-quality zoom capability. "
                                    "7. Ensure the product looks professional, clean, and appealing for online sales. "
                                    "8. Keep the product realistic and true to its original appearance. "
                                    "The final image must be optimized for PrestaShop product listings with consistent quality.")
                            
                            try:
                                enhance_result = enhance_image_with_gemini(image_search_result['image_data'], prompt, api_key)
                                if enhance_result['success']:
                                    logger.info(f"  ✓ Imagen mejorada con IA para {ean}")
                                    # Remover fondo con rembg
                                    logger.info(f"  🎨 Removiendo fondo para {ean}")
                                    try:
                                        remove_bg_result = remove_white_background(enhance_result['image_data'])
                                        if remove_bg_result['success']:
                                            image_data_final = remove_bg_result['image_data']
                                        else:
                                            image_data_final = enhance_result['image_data']
                                    except Exception as e:
                                        logger.error(f"  ❌ Error en rembg para {ean}: {e}")
                                        image_data_final = enhance_result['image_data']
                                    
                                    # Guardar imagen - Solo EAN como nombre
                                    image_filename = f"{ean}.png"
                                    
                                    images_data.append({
                                        'filename': image_filename,
                                        'data': image_data_final
                                    })
                                    logger.info(f"  ✓ Imagen guardada: {image_filename}")
                                    yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': True, 'message': 'Imagen procesada correctamente'})}\n\n"
                                else:
                                    logger.warning(f"  ⚠️ Error mejorando imagen: {enhance_result.get('error', 'Unknown')}")
                                    yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': False, 'message': 'Error mejorando imagen'})}\n\n"
                            except Exception as e:
                                logger.error(f"  ❌ Excepción en mejora de imagen para {ean}: {e}")
                                yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': False, 'message': f'Error: {str(e)}'})}\n\n"
                        else:
                            # Si no hay API key, usar imagen original - Solo EAN como nombre
                            image_filename = f"{ean}.png"
                            
                            images_data.append({
                                'filename': image_filename,
                                'data': image_search_result['image_data']
                            })
                            logger.info(f"  ✓ Imagen guardada (sin IA): {image_filename}")
                            yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': True, 'message': 'Imagen guardada (sin IA)'})}\n\n"
                    else:
                        logger.warning(f"  ⚠️ No se pudo encontrar imagen: {image_search_result.get('error', 'Unknown')}")
                        yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': False, 'message': 'No se encontró imagen'})}\n\n"
                        
                except Exception as e:
                    logger.error(f"  ❌ Error procesando imagen para {ean}: {e}")
                    yield f"data: {json.dumps({'type': 'progress', 'ean': ean, 'success': False, 'message': f'Error: {str(e)}'})}\n\n"
            
            # Crear archivo ZIP solo con imágenes
            logger.info(f"📦 Creando ZIP final con {len(images_data)} imágenes")
            if images_data:
                try:
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        # Agregar solo imágenes
                        logger.info(f"  🖼️ Agregando {len(images_data)} imágenes...")
                        for img in images_data:
                            zip_file.writestr(f"imagenes/{img['filename']}", base64.b64decode(img['data']))
                        logger.info("  ✓ Imágenes agregadas")
                        
                        # Listar contenido del ZIP
                        zip_contents = zip_file.namelist()
                        logger.info(f"  📋 Contenido del ZIP: {zip_contents}")
                    
                    zip_data = zip_buffer.getvalue()
                    
                    # Guardar en archivo temporal
                    timestamp = int(time.time() * 1000)
                    zip_filename = f"imagenes_{timestamp}.zip"
                    zip_path = os.path.join(tempfile.gettempdir(), zip_filename)
                    
                    with open(zip_path, 'wb') as f:
                        f.write(zip_data)
                    
                    logger.info(f"✅ ZIP guardado en {zip_path} ({len(zip_data)} bytes)")
                    
                    # Enviar señal de completado con nombre del archivo
                    yield f"data: {json.dumps({'type': 'complete', 'zip_filename': zip_filename})}\n\n"
                except Exception as e:
                    logger.error(f"❌ Error creando ZIP: {e}", exc_info=True)
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Error creando ZIP: {str(e)}'})}\n\n"
            else:
                logger.warning("⚠️ No hay imágenes para procesar")
                yield f"data: {json.dumps({'type': 'error', 'message': 'No se pudieron procesar imágenes'})}\n\n"
        
        except Exception as e:
            logger.error(f"❌ ERROR FATAL en process_images_only: {e}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'message': f'Error: {str(e)}'})}\n\n"
    
    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/download_zip/<filename>')
def download_zip(filename):
    """Descarga un archivo ZIP temporal"""
    try:
        # Seguridad: validar que el nombre de archivo sea válido
        if not filename.endswith('.zip') or '..' in filename or '/' in filename:
            return jsonify({'error': 'Nombre de archivo inválido'}), 400
        
        zip_path = os.path.join(tempfile.gettempdir(), filename)
        
        if not os.path.exists(zip_path):
            return jsonify({'error': 'Archivo no encontrado'}), 404
        
        # Enviar archivo y eliminarlo después
        response = send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
        # Programar eliminación del archivo después de enviarlo
        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(zip_path):
                    os.remove(zip_path)
                    logger.info(f"🗑️ Archivo temporal eliminado: {zip_path}")
            except Exception as e:
                logger.error(f"Error eliminando archivo temporal: {e}")
        
        return response
    
    except Exception as e:
        logger.error(f"Error descargando ZIP: {e}")
        return jsonify({'error': str(e)}), 500

logger.info("✅ Aplicación Flask completamente cargada y lista!")
logger.info(f"📊 Rutas registradas: {len(app.url_map._rules)}")

if __name__ == '__main__':
    logger.info("🚀 Iniciando servidor de desarrollo...")
    app.run(debug=True, host='0.0.0.0', port=5000)

