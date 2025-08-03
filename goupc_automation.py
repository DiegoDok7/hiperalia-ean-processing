import urllib.request
import json
import os
import time
from urllib.request import Request, urlopen
from typing import Optional, Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Cargar variables de entorno
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("✅ Variables de entorno cargadas desde .env")
except ImportError:
    print("⚠️  python-dotenv no instalado. Usando variables del sistema.")

# Obtener API keys
GO_UPC_API_KEY = "8c6215de81c44ee7546d8140fd980c9ba12f19f101184f383f1bbea3f01b5ea8"
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')

print("🔑 GO-UPC API key configurada")

if OPENAI_API_KEY and OPENAI_API_KEY != "xxx":
    print("🤖 OpenAI API key detectada")
elif OPENAI_API_KEY == "xxx":
    print("⚠️  Usando OpenAI API key de ejemplo. Configura tu API key real en .env")
else:
    print("❌ OpenAI API key no configurada")

# === 1. Función para consultar un código individual ===
def obtener_info_producto_goupc(codigo: str) -> Dict[str, Any]:
    """
    Consulta la API de GO-UPC para obtener información de un producto
    
    Args:
        codigo: Código UPC/EAN del producto
        
    Returns:
        Diccionario con información del producto
    """
    try:
        print(f"🔍 Consultando GO-UPC para código: {codigo}")
        
        # Verificar que tenemos API key
        if not GO_UPC_API_KEY:
            return {
                'codigo': codigo,
                'error': 'API key de GO-UPC no configurada',
                'producto_encontrado': False
            }
        
        # Crear request
        url = f'https://go-upc.com/api/v1/code/{codigo}'
        req = Request(url)
        req.add_header('Authorization', f'Bearer {GO_UPC_API_KEY}')
        
        # Hacer consulta
        try:
            response = urlopen(req)
            content = response.read()
            data = json.loads(content.decode())
            
            # Verificar si encontró el producto
            if 'product' in data and data['product']:
                product = data['product']
                
                print(f"✅ Producto encontrado: {product.get('name', 'Sin nombre')}")
                
                # Extraer información de specs si existe
                specs_dict = {}
                if 'specs' in product and product['specs']:
                    specs_dict = {spec[0]: spec[1] for spec in product['specs'] if len(spec) >= 2}
                
                # Estructurar respuesta
                resultado = {
                    'codigo': codigo,
                    'codigo_tipo': data.get('codeType', ''),
                    'nombre_producto': product.get('name', ''),
                    'descripcion': product.get('description', ''),
                    'imagen_url': product.get('imageUrl', ''),
                    'marca': product.get('brand', ''),
                    'categoria': product.get('category', ''),
                    'categoria_path': ' > '.join(product.get('categoryPath', [])),
                    'ingredientes': product.get('ingredients', {}).get('text', ''),
                    'upc': product.get('upc', ''),
                    'ean': product.get('ean', ''),
                    'barcode_url': data.get('barcodeUrl', ''),
                    'specs': specs_dict,
                    'departamento': specs_dict.get('Department', ''),
                    'producto_tipo': specs_dict.get('Commodity', ''),
                    'alergenos': specs_dict.get('Allergens', ''),
                    'organico': specs_dict.get('Organic', ''),
                    'no_gmo': specs_dict.get('Non-GMO', ''),
                    'dimensiones': {
                        'altura': specs_dict.get('Height', ''),
                        'ancho': specs_dict.get('Width', ''),
                        'largo': specs_dict.get('Length', '')
                    },
                    'producto_encontrado': True,
                    'api_usado': 'GO-UPC',
                    'intentos': 1
                }
                
                return resultado
                
            else:
                print(f"⚠️  Producto no encontrado para código: {codigo}")
                return {
                    'codigo': codigo,
                    'nombre_producto': 'PRODUCTO NO ENCONTRADO',
                    'descripcion': 'PRODUCTO NO ENCONTRADO',
                    'marca': 'PRODUCTO NO ENCONTRADO',
                    'categoria': 'PRODUCTO NO ENCONTRADO',
                    'error': 'Producto no encontrado en GO-UPC',
                    'producto_encontrado': False,
                    'api_usado': 'GO-UPC',
                    'intentos': 1
                }
                
        except urllib.error.HTTPError as e:
            error_msg = f"Error HTTP {e.code}: {e.reason}"
            print(f"❌ {error_msg}")
            return {
                'codigo': codigo,
                'error': error_msg,
                'producto_encontrado': False,
                'api_usado': 'GO-UPC',
                'intentos': 1,
                'reintentar': e.code in [400, 429]  # Bad Request o Too Many Requests
            }
            
        except Exception as e:
            error_msg = f"Error en consulta: {str(e)}"
            print(f"❌ {error_msg}")
            return {
                'codigo': codigo,
                'error': error_msg,
                'producto_encontrado': False,
                'api_usado': 'GO-UPC',
                'intentos': 1,
                'reintentar': False
            }
            
    except Exception as e:
        print(f"❌ Error general con código {codigo}: {e}")
        return {
            'codigo': codigo,
            'error': f'Error general: {str(e)}',
            'producto_encontrado': False,
            'api_usado': 'GO-UPC',
            'intentos': 1,
            'reintentar': False
        }

# === 2. Función para completar datos con IA ===
def completar_datos_con_ia(producto: Dict[str, Any]) -> Dict[str, Any]:
    """
    Completa campos vacíos de un producto usando OpenAI
    
    Args:
        producto: Diccionario con información del producto
        
    Returns:
        Producto con datos completados
    """
    if not OPENAI_API_KEY or OPENAI_API_KEY == "xxx":
        print("⚠️  OpenAI API key no configurada. Saltando completado con IA.")
        return producto
    
    # Verificar que el producto fue encontrado y tiene información básica
    if not producto.get('producto_encontrado'):
        return producto
    
    nombre = producto.get('nombre_producto', '')
    descripcion = producto.get('descripcion', '')
    
    if not nombre and not descripcion:
        return producto
    
    print(f"🤖 Completando datos con IA para: {nombre or descripcion[:50]}...")
    
    try:
        import openai
        
        # Configurar cliente OpenAI
        client = openai.OpenAI(api_key=OPENAI_API_KEY)
        
        # Preparar prompt
        prompt = f"""
        Analiza la siguiente información de un producto y completa los campos vacíos de manera profesional y precisa.
        
        Información disponible:
        - Nombre: {nombre}
        - Descripción: {descripcion}
        - Marca: {producto.get('marca', '')}
        - Categoría: {producto.get('categoria', '')}
        - Código EAN: {producto.get('codigo', '')}
        
        Campos a completar (si están vacíos):
        - marca: Marca del producto
        - categoria: Categoría principal del producto
        - departamento: Departamento de venta (ej: Alimentos, Bebidas, Limpieza, etc.)
        - producto_tipo: Tipo específico de producto
        - ingredientes: Lista de ingredientes principales (si aplica)
        - alergenos: Alérgenos comunes (si aplica)
        - organico: Si es orgánico (Sí/No)
        - no_gmo: Si es No-GMO (Sí/No)
        
        Responde SOLO en formato JSON con los campos completados. Si no puedes determinar un valor, déjalo vacío.
        """
        
        # Hacer consulta a OpenAI
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Eres un experto en análisis de productos. Responde solo en formato JSON válido."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=500
        )
        
        # Procesar respuesta
        try:
            respuesta_ia = json.loads(response.choices[0].message.content)
            
            # Actualizar campos vacíos con datos de IA
            campos_ia = ['marca', 'categoria', 'departamento', 'producto_tipo', 'ingredientes', 'alergenos', 'organico', 'no_gmo']
            
            for campo in campos_ia:
                if campo in respuesta_ia and respuesta_ia[campo]:
                    valor_actual = producto.get(campo, '')
                    if not valor_actual or valor_actual in ['N/A', 'PRODUCTO NO ENCONTRADO']:
                        producto[campo] = respuesta_ia[campo]
                        print(f"   ✅ {campo}: {respuesta_ia[campo]}")
            
            producto['datos_completados_ia'] = True
            
        except json.JSONDecodeError:
            print("   ⚠️  Error al parsear respuesta de IA")
            producto['datos_completados_ia'] = False
            
    except Exception as e:
        print(f"   ❌ Error en completado con IA: {e}")
        producto['datos_completados_ia'] = False
    
    return producto

# === 3. Función para procesar con reintentos ===
def procesar_codigos_con_reintentos(codigos: list, max_intentos: int = 2) -> list:
    """
    Procesa códigos con sistema de reintentos para errores de API
    
    Args:
        codigos: Lista de códigos UPC/EAN
        max_intentos: Número máximo de intentos por código
        
    Returns:
        Lista de resultados procesados
    """
    print(f"🚀 Procesando {len(codigos)} códigos con sistema de reintentos...")
    resultados = []
    
    # Primer procesamiento
    for i, codigo in enumerate(codigos, 1):
        if codigo.strip():
            print(f"[{i}/{len(codigos)}] Procesando código: {codigo.strip()}")
            resultado = obtener_info_producto_goupc(codigo.strip())
            resultados.append(resultado)
    
    # Identificar productos que necesitan reintento
    productos_reintentar = [r for r in resultados if r.get('reintentar', False)]
    
    if productos_reintentar:
        print(f"\n⏳ {len(productos_reintentar)} productos con errores de API. Esperando 60 segundos...")
        time.sleep(60)
        
        print("🔄 Reintentando productos con errores...")
        for producto in productos_reintentar:
            codigo = producto['codigo']
            print(f"🔄 Reintentando: {codigo}")
            
            nuevo_resultado = obtener_info_producto_goupc(codigo)
            nuevo_resultado['intentos'] = 2
            
            # Actualizar el resultado original
            for i, r in enumerate(resultados):
                if r['codigo'] == codigo:
                    resultados[i] = nuevo_resultado
                    break
    
    # Completar datos con IA para productos encontrados
    print(f"\n🤖 Completando datos con IA...")
    productos_para_ia = [r for r in resultados if r.get('producto_encontrado') and 
                        (r.get('nombre_producto') or r.get('descripcion'))]
    
    for i, producto in enumerate(productos_para_ia, 1):
        print(f"[{i}/{len(productos_para_ia)}] Completando con IA...")
        resultados[resultados.index(producto)] = completar_datos_con_ia(producto)
    
    # Estadísticas finales
    encontrados = len([r for r in resultados if r.get('producto_encontrado')])
    completados_ia = len([r for r in resultados if r.get('datos_completados_ia')])
    
    print(f"\n📊 ESTADÍSTICAS FINALES:")
    print(f"   • Total procesados: {len(resultados)}")
    print(f"   • Productos encontrados: {encontrados}")
    print(f"   • Completados con IA: {completados_ia}")
    print(f"   • Tasa de éxito: {(encontrados/len(resultados)*100):.1f}%")
    
    return resultados

# === 4. Función para leer códigos desde archivo ===
def leer_codigos_desde_archivo(ruta_archivo: str) -> list:
    """
    Lee códigos UPC/EAN desde un archivo de texto
    
    Args:
        ruta_archivo: Ruta al archivo con códigos
        
    Returns:
        Lista de códigos
    """
    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as file:
            contenido = file.read().strip()
            codigos = [codigo.strip() for codigo in contenido.split('\n') if codigo.strip()]
            print(f"📁 Leídos {len(codigos)} códigos desde {ruta_archivo}")
            return codigos
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {ruta_archivo}")
        return []
    except Exception as e:
        print(f"❌ Error al leer el archivo: {e}")
        return []

# === 5. Función para procesar múltiples códigos (legacy) ===
def procesar_codigos_goupc(codigos: list) -> list:
    """
    Procesa una lista de códigos con GO-UPC (versión legacy)
    
    Args:
        codigos: Lista de códigos UPC/EAN
        
    Returns:
        Lista de diccionarios con información de productos
    """
    return procesar_codigos_con_reintentos(codigos)

# === 6. Función para guardar resultados en Excel ===
def guardar_excel_goupc(datos: list, ruta_archivo: str):
    """
    Guarda los datos en formato Excel con formato profesional
    """
    if not datos:
        print("No hay datos para guardar")
        return
    
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos GO-UPC"
        
        # Definir columnas principales
        columnas_principales = [
            'codigo', 'codigo_tipo', 'nombre_producto', 'descripcion', 
            'marca', 'categoria', 'categoria_path', 'departamento',
            'producto_tipo', 'imagen_url', 'upc', 'ean',
            'ingredientes', 'alergenos', 'organico', 'no_gmo',
            'altura', 'ancho', 'largo', 'barcode_url',
            'producto_encontrado', 'api_usado', 'intentos', 'datos_completados_ia', 'error'
        ]
        
        # Escribir encabezados
        for col, campo in enumerate(columnas_principales, 1):
            cell = ws.cell(row=1, column=col, value=campo.replace('_', ' ').title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir datos
        for row, producto in enumerate(datos, 2):
            for col, campo in enumerate(columnas_principales, 1):
                if campo in ['altura', 'ancho', 'largo']:
                    # Extraer dimensiones
                    valor = producto.get('dimensiones', {}).get(campo, '')
                else:
                    valor = producto.get(campo, '')
                
                # Limpiar y formatear valor
                if isinstance(valor, str):
                    valor = valor.replace('\n', ' ').replace('\r', ' ').strip()
                elif isinstance(valor, bool):
                    valor = 'Sí' if valor else 'No'
                elif valor is None:
                    valor = ''
                
                ws.cell(row=row, column=col, value=valor)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Añadir hoja de especificaciones detalladas
        ws_specs = wb.create_sheet("Especificaciones")
        specs_row = 1
        
        # Encabezados de specs
        ws_specs.cell(row=specs_row, column=1, value="Código").font = Font(bold=True)
        ws_specs.cell(row=specs_row, column=2, value="Producto").font = Font(bold=True)
        ws_specs.cell(row=specs_row, column=3, value="Especificación").font = Font(bold=True)
        ws_specs.cell(row=specs_row, column=4, value="Valor").font = Font(bold=True)
        specs_row += 1
        
        # Datos de specs
        for producto in datos:
            if producto.get('specs') and isinstance(producto['specs'], dict):
                for spec_name, spec_value in producto['specs'].items():
                    ws_specs.cell(row=specs_row, column=1, value=producto.get('codigo', ''))
                    ws_specs.cell(row=specs_row, column=2, value=producto.get('nombre_producto', ''))
                    ws_specs.cell(row=specs_row, column=3, value=spec_name)
                    ws_specs.cell(row=specs_row, column=4, value=spec_value)
                    specs_row += 1
        
        # Guardar archivo
        wb.save(ruta_archivo)
        print(f"✅ Archivo Excel guardado: {ruta_archivo}")
        print(f"📊 {len(datos)} productos procesados")
        print(f"📋 {len(columnas_principales)} columnas de datos")
        
    except Exception as e:
        print(f"❌ Error al guardar Excel: {e}")

# === 7. Función para guardar resultados en JSON ===
def guardar_json(datos: list, ruta_archivo: str):
    """
    Guarda los datos en formato JSON
    """
    try:
        with open(ruta_archivo, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        print(f"✅ Archivo JSON guardado: {ruta_archivo}")
    except Exception as e:
        print(f"❌ Error al guardar JSON: {e}")

# === 8. Función para mostrar información de un producto ===
def mostrar_producto(producto: Dict[str, Any]):
    """
    Muestra información detallada de un producto
    """
    print("\n" + "="*50)
    print(f"📦 PRODUCTO: {producto.get('nombre_producto', 'N/A')}")
    print("="*50)
    print(f"🔍 Código: {producto.get('codigo', 'N/A')}")
    print(f"🏷️  Marca: {producto.get('marca', 'N/A')}")
    print(f"📂 Categoría: {producto.get('categoria', 'N/A')}")
    print(f"📝 Descripción: {producto.get('descripcion', 'N/A')}")
    
    if producto.get('ingredientes'):
        print(f"🧪 Ingredientes: {producto['ingredientes'][:100]}...")
    
    if producto.get('alergenos'):
        print(f"⚠️  Alérgenos: {producto['alergenos']}")
    
    if producto.get('imagen_url'):
        print(f"🖼️  Imagen: {producto['imagen_url']}")
    
    print(f"✅ Encontrado: {'Sí' if producto.get('producto_encontrado') else 'No'}")
    print(f"🤖 IA: {'Sí' if producto.get('datos_completados_ia') else 'No'}")
    print(f"🔄 Intentos: {producto.get('intentos', 1)}")
    print("="*50)

# === 9. Función de prueba rápida ===
def prueba_rapida():
    """
    Función de prueba rápida con un código de ejemplo
    """
    print("🧪 PRUEBA RÁPIDA GO-UPC CON IA")
    print("-" * 40)
    
    # Código de ejemplo del documentation
    codigo_prueba = "781138811156"
    
    resultado = obtener_info_producto_goupc(codigo_prueba)
    resultado = completar_datos_con_ia(resultado)
    mostrar_producto(resultado)
    
    return resultado

# === Ejemplo de uso ===
if __name__ == "__main__":
    print("🚀 GO-UPC Automation con IA")
    print("=" * 40)
    
    # Verificar configuración
    if not GO_UPC_API_KEY:
        print("❌ Configura GO_UPC_API_KEY en tu archivo .env")
        exit(1)
    
    # Ejecutar prueba rápida
    prueba_rapida()
    
    print("\n💡 Ejemplos de uso:")
    print("# Procesar con reintentos y IA:")
    print("# codigos = ['781138811156', '843248155091']")
    print("# resultados = procesar_codigos_con_reintentos(codigos)")
    print("# guardar_json(resultados, 'productos_goupc_ia.json')") 